import time
import math
import cv2
from config import *
from ultralytics import YOLO
from comm import LoRaComm
import subprocess
import threading
from threading import Lock
import queue
import numpy as np
import os
import RPi.GPIO as GPIO
from openpyxl import Workbook, load_workbook


# ================= CPU THREAD LIMIT =================
# os.environ["OMP_NUM_THREADS"] = "4"
# os.environ["OPENBLAS_NUM_THREADS"] = "4"
# os.environ["MKL_NUM_THREADS"] = "4"
# os.environ["NUMEXPR_NUM_THREADS"] = "4"


# ================= GPIO =================
LED_PIN = 24
GPIO.setmode(GPIO.BCM)
GPIO.setup(LED_PIN, GPIO.OUT)


# ================= MODEL =================
# model = YOLO("best_ncnn_model", task="classify")
model = YOLO("best.onnx", task="classify")
model.overrides["batch"] = 8
model.overrides["device"] = "cpu"

comm = LoRaComm()
lora_tx_lock = Lock()

tx_buffer = []
last_tx_time = 0

prev_time = time.time()
fps = 0.0


# ================= EXCEL LOG =================
EXCEL_PATH = "plant_data.xlsx"

if os.path.exists(EXCEL_PATH):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Latitude", "Longitude", "Cell", "Disease", "Infected Area", "Healthy Area", "Frames"])


def log_cell_to_excel(cell, data):
    lat, lon = data["gps"]
    ts = time.strftime("%Y-%m-%d %H:%M:%S")

    if not data["diseases"]:
        ws.append([ts, lat, lon, str(cell), "NO DATA", 0, 0, 0])
    else:
        for disease, s in data["diseases"].items():
            ws.append([
                ts, lat, lon, str(cell),
                disease,
                int(s["infected_area"] / px_to_area_scale),
                int(s["healthy_area"] / px_to_area_scale),
                s["frames"]
            ])

    wb.save(EXCEL_PATH)
    
# ================= FAKE GPS =================
# GRID SCAN STATE (S-PATTERN)
last_gps_update_time = 0
scan_gx = 0
scan_gy = 0
scan_dir = 1   # +1 = left→right, -1 = right→left
scan_done = False

def fake_gps():
    global scan_gx, scan_gy, scan_dir, scan_done
    global drone_lat, drone_lon, last_gps_update_time

    if scan_done:
        return drone_lat, drone_lon

    now = time.time()
    # ONLY advance the grid if the interval has passed
    if now - last_gps_update_time >= GPS_UPDATE_INTERVAL:
        
        drone_lat, drone_lon = grid_to_gps(scan_gx, scan_gy)
        scan_gx += scan_dir
        
        if scan_gx >= GRID_COLS or scan_gx < 0:
            scan_dir *= -1
            scan_gx += scan_dir
            scan_gy += 1
            
            if scan_gy >= GRID_ROWS:
                scan_done = True
                scan_gx += 1
                print("[SCAN]: Grid coverage complete")
        
        last_gps_update_time = now

    return drone_lat, drone_lon


# ================= CAMERA =================
frame_q = queue.Queue(maxsize=1)
stop_event = threading.Event()

if LIVE_STREAM:
    cam_proc = subprocess.Popen(
        [
            "rpicam-vid",
            "--width", str(WIDTH),
            "--height", str(HEIGHT),
            "--framerate", "10",
            "--codec", "yuv420",
            "--nopreview",
            "-t", "0",
            "-o", "-"
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL
    )

    def cam_reader():
        while not stop_event.is_set():
            raw = cam_proc.stdout.read(FRAME_SIZE)
            if len(raw) != FRAME_SIZE:
                continue

            yuv = np.frombuffer(raw, np.uint8).reshape((int(HEIGHT * 1.5), WIDTH))
            frame = cv2.cvtColor(yuv, cv2.COLOR_YUV2BGR_I420)

            if frame_q.full():
                frame_q.get_nowait()

            frame_q.put(frame)

    threading.Thread(target=cam_reader, daemon=True).start()

else:
    cap = cv2.VideoCapture("./apple.mp4")


# ================= GRID DATA =================
grid_data = {}
current_cell = None


# ================= GRID UPDATE =================
def update_grid(cell, disease, infected_px, healthy_px):

    grid = grid_data.setdefault(cell, {
        "gps": grid_to_gps(*cell),
        "diseases": {}
    })

    d = grid["diseases"].setdefault(disease, {
        "infected_area": 0,
        "healthy_area": 0,
        "frames": 0
    })

    d["infected_area"] += infected_px
    d["healthy_area"] += healthy_px


# ================= GPS HELPERS =================
def gps_to_xy(lat, lon):
    dlat = math.radians(lat - LAT0)
    dlon = math.radians(lon - LON0)
    x = EARTH_RADIUS * dlon * math.cos(math.radians(LAT0))
    y = EARTH_RADIUS * dlat
    return x, y

def gps_to_grid(lat, lon):
    x, y = gps_to_xy(lat, lon)
    if x < 0 or y < 0 or x > FIELD_WIDTH_M or y > FIELD_HEIGHT_M:
        return (-1, -1)
    return int(x // CELL_W), int(y // CELL_H)


def grid_to_gps(gx, gy):
    x = (gx + 0.5) * CELL_W
    y = (gy + 0.5) * CELL_H
    lat = LAT0 + math.degrees(y / EARTH_RADIUS)
    lon = LON0 + math.degrees(x / (EARTH_RADIUS * math.cos(math.radians(LAT0))))
    return lat, lon


# ================= TRANSMISSION =================
MSG_ID = 1

def save_grid(cell, data):
    global MSG_ID

    lat, lon = data["gps"]
    gx, gy = cell

    payload = f"[ID]:{MSG_ID} | CELL {gx, gy} | GPS {int(lat*gps_scale)},{int(lon*gps_scale)}"

    if data["diseases"]:
        for disease, s in data["diseases"].items():
            payload += f" | {disease}:{int(s['infected_area']/px_to_area_scale)}:{int(s['healthy_area']/px_to_area_scale)}:{s['frames']}"
    else:
        payload += " | NO DATA"

    payload += "\n"

    tx_buffer.append(payload)
    MSG_ID += 1


def try_transmit():
    global last_tx_time

    if not tx_buffer:
        return

    now = time.time()

    if now - last_tx_time > TX_INTERVAL:
        payload = "".join(tx_buffer)

        with lora_tx_lock:
            comm.encrypt_and_send(payload)
            tx_buffer.clear()

        last_tx_time = now


# ================= LED =================
def blink_led(times=1, duration=0.2):
    for _ in range(times):
        GPIO.output(LED_PIN, GPIO.HIGH)
        time.sleep(duration)
        GPIO.output(LED_PIN, GPIO.LOW)
        time.sleep(duration)


# ================= By Parts classification =================
def tiled_classification(frame, tile_size=240, prob_thresh=0.75):
    h, w = frame.shape[:2]

    tiles = []
    tile_meta = []

    # -------- Collect Tiles --------
    for y in range(0, h, tile_size):
        for x in range(0, w, tile_size):

            tile = frame[y:y+tile_size, x:x+tile_size]

            if tile.size == 0:
                continue

            tile_h, tile_w = tile.shape[:2]
            real_area = tile_h * tile_w

            tile_resized = cv2.resize(tile, (INFERENCE_SIZE, INFERENCE_SIZE))

            tiles.append(tile_resized)
            tile_meta.append(real_area)

    if not tiles:
        return {}

    # 🔥 IMPORTANT FIX: convert list -> numpy batch
    batch = np.stack(tiles, axis=0)

    batch = np.stack(tiles, axis=0).astype(np.float32) / 255.0
    
    # -------- Batched Inference --------
    results = model(batch, verbose=False)

    disease_stats = {}

    # -------- Process Results --------
    for i, r in enumerate(results):

        if r.probs is None:
            continue

        conf = float(r.probs.top1conf)
        cls_id = int(r.probs.top1)

        if conf < prob_thresh:
            continue

        disease = r.names[cls_id]
        real_area = tile_meta[i]

        infected_px = real_area * conf
        healthy_px = real_area * (1 - conf)

        d = disease_stats.setdefault(disease, {
            "infected_area": 0,
            "healthy_area": 0,
            "frames": 0
        })

        d["infected_area"] += infected_px
        d["healthy_area"] += healthy_px

    return disease_stats



# ================= MAIN LOOP =================
print("\n=== Drone Disease Classification Started ===\n")

try:
    while True:

        # ----- FRAME SOURCE -----
        if LIVE_STREAM:
            if frame_q.empty():
                continue
            frame = frame_q.get()
        else:
            ret, frame = cap.read()
            if not ret:
                break

        # ----- GPS -----
        lat, lon = fake_gps()
        cell = gps_to_grid(lat, lon)

        if cell == (-1, -1):
            continue

        # ----- INFERENCE -----
        y_offset = 120
        
        diseases = tiled_classification(frame)
        
        for disease, stats in diseases.items():
            update_grid(cell, disease, stats["infected_area"], stats["healthy_area"])
            
            total_area = stats["infected_area"] + stats["healthy_area"]
            if total_area > 0:
                avg_conf = stats["infected_area"] / total_area

                cv2.putText(frame, f"{disease} ({avg_conf*100:.1f}%)", (10, y_offset), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
                y_offset += 25

   
        # ----- CELL TRANSITION -----
        if current_cell is None:
            current_cell = cell

        elif cell != current_cell:
            data = grid_data.pop(
                current_cell,
                {"gps": grid_to_gps(*current_cell), "diseases": {}}
            )

            save_grid(current_cell, data)
            log_cell_to_excel(current_cell, data)

            threading.Thread(target=blink_led, args=(2,), daemon=True).start()

            current_cell = cell

        # ----- FPS -----
        now = time.time()
        fps = 0.9 * fps + 0.1 * (1 / (now - prev_time))
        prev_time = now

        cv2.putText(frame, f"FPS:{fps:.2f}", (10, 35), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,255,0), 2)

        cv2.imshow("Drone Feed", frame)

        # ----- TRANSMIT -----
        try_transmit()

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
finally:
    GPIO.output(LED_PIN, GPIO.LOW)
    GPIO.cleanup()
    cv2.destroyAllWindows()
