import time
import random
import math
import cv2
from config import *
from ultralytics import YOLO
from comm import LoRaComm
import subprocess
import threading
from threading import Lock
import queue
import numpy as np
import os
import RPi.GPIO as GPIO
from openpyxl import Workbook, load_workbook


# =================== SYSTEM THREAD LIMITS ========================
os.environ["OMP_NUM_THREADS"] = "2"
os.environ["OPENBLAS_NUM_THREADS"] = "2"
os.environ["MKL_NUM_THREADS"] = "2"
os.environ["NUMEXPR_NUM_THREADS"] = "2"
os.environ["VECLIB_MAXIMUM_THREADS"] = "2"
os.environ["NUMBA_NUM_THREADS"] = "2"


# ============================ GPIO INIT ============================
LED_PIN = 24  # GPIO pin your LED is connected to
GPIO.setmode(GPIO.BCM)
GPIO.setup(LED_PIN, GPIO.OUT)


# ============================== MODEL + COMM INIT ===================
# yolo export model=best.pt format=ncnn imgsz=320 half=True
model = YOLO("best_ncnn_model", task="classify")

# Lora Comm Init
comm = LoRaComm()
lora_tx_lock = Lock()

# TX Buffer
tx_buffer = []
last_tx_time = 0

# FPS Calculation
fps_counter = 0
start_time = time.time()


# ============================== DATA LOG INIT ========================
EXCEL_PATH = "plant_data.xlsx"  # Excel file path

if os.path.exists(EXCEL_PATH):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Latitude", "Longitude", "Cell", "Disease", "Infected Area", "Healthy Area", "Leaves"])

def log_cell_to_excel(cell, data):
    lat, lon = data["gps"]
    ts = time.strftime("%Y-%m-%d %H:%M:%S")

    if not data["diseases"]:
        ws.append([
            ts, lat, lon,
            f"{cell}",
            "NO DATA",
            0, 0, 0
        ])
    else:
        for disease, s in data["diseases"].items():
            ws.append([ts, lat, lon, f"{cell}", disease, int(s["infected_area"] / px_to_area_scale), 
                       int(s["healthy_area"] / px_to_area_scale), s["frames"]])
    
    wb.save(EXCEL_PATH)


# ============================== CAMERA INIT ==============================
frame_q = queue.Queue(maxsize=1)
stop_event = threading.Event()

if LIVE_STREAM:
    print("[CAM]: Using Pi Camera (rpicam-vid)")

    cam_proc = subprocess.Popen(
        [
            "rpicam-vid",
            "--width", str(WIDTH),
            "--height", str(HEIGHT),
            "--framerate", "10",
            "--codec", "yuv420",
            "--nopreview",
            "-t", "0",
            "-o", "-"
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL
    )

    def cam_reader():
        while not stop_event.is_set():
            raw = cam_proc.stdout.read(FRAME_SIZE)
            if len(raw) != FRAME_SIZE:
                continue
            try:
                yuv = np.frombuffer(raw, np.uint8).reshape((int(HEIGHT * 1.5), WIDTH))
                frame = cv2.cvtColor(yuv, cv2.COLOR_YUV2BGR_I420)
            except:
                time.sleep(0.5)
                continue

            if frame_q.full():
                frame_q.get_nowait()
            frame_q.put(frame)

    threading.Thread(target=cam_reader, daemon=True).start()

else:
    print("[CAM]: Using video file")
    cap = cv2.VideoCapture("./apple.mp4")


# ============================ GPS SIMULATION =============================
# GRID SCAN STATE (S-PATTERN)
last_gps_update_time = 0
scan_gx = 0
scan_gy = 0
scan_dir = 1   # +1 = left→right, -1 = right→left
scan_done = False

def fake_gps():
    global scan_gx, scan_gy, scan_dir, scan_done
    global drone_lat, drone_lon, last_gps_update_time

    if scan_done:
        return drone_lat, drone_lon

    now = time.time()
    if now - last_gps_update_time >= GPS_UPDATE_INTERVAL:     # ONLY advance the grid if the interval has passed
        
        drone_lat, drone_lon = grid_to_gps(scan_gx, scan_gy)
        scan_gx += scan_dir

        # End of row reached
        if scan_gx >= GRID_COLS or scan_gx < 0:
            scan_dir *= -1
            scan_gx += scan_dir
            scan_gy += 1

            if scan_gy >= GRID_ROWS:
                scan_done = True
                scan_gx += 1
                print("[SCAN]: Grid coverage complete")
        
        last_gps_update_time = now # Reset the timer

    return drone_lat, drone_lon

# ===================== GRID / GPS UTILITIES ==============================
def gps_to_xy(lat, lon):
    dlat = math.radians(lat - LAT0)
    dlon = math.radians(lon - LON0)
    x = EARTH_RADIUS * dlon * math.cos(math.radians(LAT0))
    y = EARTH_RADIUS * dlat
    return x, y

def gps_to_grid(lat, lon):
    x, y = gps_to_xy(lat, lon)
    if x < 0 or y < 0 or x > FIELD_WIDTH_M or y > FIELD_HEIGHT_M:
        return (-1, -1)
    return int(x // CELL_W), int(y // CELL_H)

def grid_to_gps(gx, gy):
    x = (gx + 0.5) * CELL_W
    y = (gy + 0.5) * CELL_H
    lat = LAT0 + math.degrees(y / EARTH_RADIUS)
    lon = LON0 + math.degrees(x / (EARTH_RADIUS * math.cos(math.radians(LAT0))))
    return lat, lon


# ======================= YOLO CLS==================================
def yolo_cls_infer(frame, prob_thresh=0.50, topk=2):

    results = model.predict(frame, imgsz=INFERENCE_SIZE, verbose=False)
    r = results[0]
    
    speed = r.speed
    print(f"Inference Time:   {speed['inference']:.2f}ms")
    
    if r.probs is None:
        return []

    probs = r.probs.data.cpu().numpy()

    # -------- TOP K CLASSES --------
    top_ids = probs.argsort()[-topk:][::-1]

    h, w = frame.shape[:2]
    area = h * w

    detections = []

    for cls_id in top_ids:

        conf = probs[cls_id]

        # skip low confidence
        # if conf < prob_thresh:
        #     continue

        disease = r.names[int(cls_id)]

        infected_px = area * conf
        healthy_px = area * (1 - conf)

        detections.append((disease, infected_px, healthy_px, conf))

    return detections



# ========================== GRID UPDATE ==========================
def update_grid(cell, disease, infected_px, healthy_px):

    grid = grid_data.setdefault(cell, {
        "gps": grid_to_gps(*cell),
        "diseases": {}
    })

    d = grid["diseases"].setdefault(disease, {
        "infected_area": 0,
        "healthy_area": 0,
        "frames": 0
    })

    d["infected_area"] += infected_px
    d["healthy_area"] += healthy_px
    d["frames"] += 1


# ========================== TRANSMISSION ===============================
# def try_transmit():
#     global last_tx_time, tx_buffer

#     if not tx_buffer:
#         return

#     now = time.time()
#     if now - last_tx_time > TX_INTERVAL:   # Send every TX_INTERVAL seconds
#         final_payload = "".join(tx_buffer)
        
#         with lora_tx_lock:
#             comm.encrypt_and_send(final_payload)
#             tx_buffer.clear()
#             last_tx_time = now
#             print("[TX]: SENT TO BASE\n")


last_tx_time = 0
last_tx_duration = 0   # store duration of previous TX

def estimate_tx_interval(payload_len):
    airtime = 0.05 * payload_len   # rough LoRa scaling (depends on SF/BW)
    decrypt_time = 0.002 * payload_len
    return airtime + decrypt_time + 0.5   # margin

def try_transmit():
    global last_tx_time, tx_buffer, last_tx_duration

    if not tx_buffer:
        return

    now = time.time()

    # use previous TX duration
    if now - last_tx_time > last_tx_duration:
        
        final_payload = "".join(tx_buffer)

        # compute NEXT duration BEFORE sending
        next_duration = estimate_tx_interval(len(final_payload))

        with lora_tx_lock:
            comm.encrypt_and_send(final_payload)
            tx_buffer.clear()

            last_tx_time = now
            last_tx_duration = next_duration   # store for next cycle

            print(f"[TX]: SENT | next wait = {next_duration:.2f}sec\n")


# ========================== Data Buffer ===============================
def save_grid(cell, data):
    global MSG_ID
    lat, lon = data["gps"]
    gx, gy = cell
    payload = f"[ID]:{MSG_ID} | CELL {gx, gy} | GPS {int(lat*gps_scale)},{int(lon*gps_scale)}"

    if data["diseases"]:
        for disease, s in data["diseases"].items():
            payload += f" | {disease} : {int(s['infected_area']/px_to_area_scale)} : {int(s['healthy_area']/px_to_area_scale)} : {s['frames']}"
    else:
        payload += " | NO DATA"
        
    payload += "\n"
    tx_buffer.append(payload)
    MSG_ID += 1
    print("[TX-BUFFER]: QUEUED")
    
    
def blink_led(times=1, duration=0.2):
    for _ in range(times):
        GPIO.output(LED_PIN, GPIO.HIGH)
        time.sleep(duration)
        GPIO.output(LED_PIN, GPIO.LOW)
        time.sleep(duration)


# =========================================================
# MAIN LOOP
# =========================================================
print("\n================= Drone Grid-Level Plant Disease Classification Started =================\n")
try:
    while True:

        # ================= FRAME SOURCE =================
        if LIVE_STREAM:
            if frame_q.empty():
                continue
            frame = frame_q.get()
        else:
            ret, frame = cap.read()
            if not ret:
                break
        # ==========================================================
        
        
        # ================= GPS & GRID ===============================
        lat, lon = fake_gps()
        cell = gps_to_grid(lat, lon)
        if cell == (-1, -1):
            continue
        
        # if scan finished, flush last cell exactly once
        if scan_done and current_cell is not None:
            data = grid_data.pop(current_cell, {"gps": grid_to_gps(*current_cell), "diseases": {}})
            print(f"[SCAN]: Final cell {current_cell} : SEND DATA")
            
            log_cell_to_excel(current_cell, data)   # Sent to Log
            save_grid(current_cell, data)           # Sent to Base
            try_transmit()                          # Sent to Base
            break
        # ================================================================
        
        
        # ========================= YOLO CLS =============================
        detections = yolo_cls_infer(frame)
       
        for i, (disease, infected, healthy, conf) in enumerate(detections): # Create Dataset of current cell
            
            cv2.putText(frame, f"Top{i+1}: {disease} ({conf*100:.1f}%)",(10, 125 + i*25),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,0,255), 2)
            
            # Noramal flow matching & updating leaf
            update_grid(cell, disease, infected, healthy)
        # ==================================================================



        # ======================== CELL TRANSITION =========================
        if current_cell is None:
            current_cell = cell

        elif cell != current_cell:
            data = grid_data.pop(
                current_cell,
                {"gps": grid_to_gps(*current_cell), "diseases": {}}
            )
            print(f"[CELL]: Leaving cell {current_cell} : SEND DATA")
            save_grid(current_cell, data)
            
            # ---------- EXCEL LOG ----------
            # log_cell_to_excel(current_cell, data)
            
            # Blink LED parallel
            threading.Thread(target=blink_led,args=(2,),daemon=True).start()
        
            current_cell = cell
        # ===================================================================
        
        
        # ================= FPS ==============================================
        fps_counter += 1
        elapsed_time = time.time() - start_time
        fps = fps_counter / elapsed_time if elapsed_time > 0 else 0


        cv2.putText(frame, f"FPS: {fps:.2f}", (10,35), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0,255,0), 2)
        
        cell_text = f"Cell: ({cell[0]}, {cell[1]})"
        cv2.putText(frame, cell_text, (10, 65), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 0, 0), 2)
    
        cv2.putText(frame, f"Classified: {len(detections)}", (10, 95), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
        
        # cv2.imshow("Drone Live Feed", frame)
        print(f"FPS: {fps:.2f}")
        # ======================================================================
        
        
        # ================= Data Transmission =================
        try_transmit()
        # =================================================== 
        
        
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break
    
finally:
    GPIO.output(LED_PIN, GPIO.LOW)
    GPIO.cleanup()
    cv2.destroyAllWindows()
    if LIVE_STREAM:
        stop_event.set()
        cam_proc.terminate()